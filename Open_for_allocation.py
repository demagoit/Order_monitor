import pandas as pd
import numpy as np

from datetime import date, datetime, timedelta

import os
from openpyxl import load_workbook

file_folder = 'Deliveries'

in_file = {
    # 'file': "Open_orders_test_12.05.2022.xlsx",
    'file': "Open_orders_test_18.08.2023.xlsx",
    'sheet_name': "Open_orders_A101",
    'stock_sheet': "MBEW_A101",
    # 'stock_sheet': "A101_Stock",
    'PL40_Prio_List': 'Prio_List_PL40'
}

out_file = {
    'file': "Open_orders_test_py_out.xlsx"
}

confirmation_file = {
    'file': "Delivery_confirmations_py_out.xlsx"
}


def read_in_file(file: str, folder=file_folder, in_SheetName='', header_row=2):
    in_file = os.path.join(folder, file)
    print(in_file)

    try:
        wb = load_workbook(filename=in_file, data_only=True)

        if in_SheetName == '':
            print("Loading last sheet in file")
            df = pd.DataFrame(wb[wb.sheetnames[-1]].values)

        else:
            df = pd.DataFrame(wb[in_SheetName].values)

        sheets = wb.sheetnames
        wb.close()

        column_names = rename_columns(df.iloc[header_row].values)
        df.columns = column_names
        df.drop(index=list(np.arange(header_row+1)), inplace=True)
        df.reset_index(drop=True, inplace=True)

    except ValueError:
        # wb = pd.read_excel(in_file, sheet_name=-1)#, header=None)

        # sheets = list(wb.keys())
        if in_SheetName == '':
            print("Loading last sheet in file")
            df = pd.read_excel(in_file, sheet_name=-1, header=header_row)
            # df = wb.get(sheets[-1])
        else:
            df = pd.read_excel(in_file, in_SheetName, header=header_row)
            # df = wb.get(in_SheetName)

        column_names = rename_columns(df.columns.values)
        df.columns = column_names
        sheets = []

        # del (wb)

    except FileNotFoundError:
        df = pd.DataFrame()
        print(f'File {in_file} not found')

    except Exception as Err:
        print(Err)
        os._exit(1)

    return df, sheets

def write_out_file(df: pd.DataFrame, file_name: str, file_folder: str = file_folder, mode: str = 'a', sheet_name: str = None, if_sheet_exists: str = 'replace',  print_msg: str = None):

    with pd.ExcelWriter(os.path.join(file_folder, file_name), mode=mode, if_sheet_exists=if_sheet_exists) as writer:
        if print_msg is None:
            if mode == 'a':
                print_msg = f"Updating {file_name} file"
            else:
                print_msg = f"Creating {file_name} file"
        if sheet_name is not None:
            print_msg += f", adding {sheet_name} sheet"

        print(print_msg, '\n')

        if sheet_name is None:
            df.to_excel(writer, index=False)
        else:
            df.to_excel(writer, sheet_name=cur_conf_col, index=False)

def rename_columns(column_names):
    for item in enumerate(column_names):
        try:
            column_names[item[0]] = '_'.join(item[1].split())
        except:
            column_names[item[0]] = item[1]
    return column_names


df, sheets = read_in_file(
    file=in_file['file'], folder=file_folder, in_SheetName=in_file['sheet_name'], header_row=2)
df_a101, _ = read_in_file(
    file=in_file['file'], folder=file_folder, in_SheetName=in_file['stock_sheet'], header_row=1)
print(sheets)

df = df.convert_dtypes()
# drop rows with all N/A values
df = df.dropna(thresh=len(df.columns))

df_a101 = df_a101.convert_dtypes()
# drop rows with all N/A values
df_a101 = df_a101.dropna(thresh=len(df_a101.columns))
df_a101 = df_a101[['Material', 'Total_Stock']]
df_a101.set_index('Material', drop=True, inplace=True)

# filtration part
df = df.drop(['Product_hierarchy', 'Product_hierarchy_text',
             'Plant', 'Order_Value'], axis=1)
df.insert(0, 'item_index', df['Sales_document'] +
          '_' + df['Sales_Document_Item'])
df.insert(1, 'SO_Mat_index', df['Sales_document'] + '_' + df['Material'])
df.drop(df[df['SD_Document_Category'] != 'Order'].index, inplace=True)

df = df.join(df_a101, on='Material', how='left')

# check changes in confirmation dates
# drop lines without confirmation date
df_cur_conf = df[df['Confirmed_Date'] != '#']

# drop already partially delivered lines
df_part_delivered = df_cur_conf[(df_cur_conf['Delivery_Quantity'] > 0) &
                                (df_cur_conf['Confirmed_Date'] <= df_cur_conf['Created_on'].max())]
df_cur_conf = df_cur_conf.drop(df_part_delivered.index)
df_cur_conf = df_cur_conf.sort_values(
    by=['item_index', 'Confirmed_Date'], ascending=[True, True])
cur_conf_col = 'Confirmation_' + df_cur_conf['Created_on'].max()

# def test(base_date, comparision_date):
#     '''
#     take:
#     base_date - requested date
#     comparision_date - confirmed date
#     returns:
#     is_overdue -
#     adj_base_date - base_date changed to nearest thursday in future
#     '''
#     base_date = pd.to_datetime(
#         base_date, yearfirst=True, errors='coerce').dt.date
#     comparision_date = pd.to_datetime(
#         comparision_date, yearfirst=True, errors='coerce').dt.date

#     adj_base_date = base_date.apply(lambda x:
#                                     4-x.isoweekday() if x.isoweekday() <= 4 else 7+4-x.isoweekday())
# #     adj_base_date = base_date + pd.Timedelta(adj_base_date, unit='d')
#     adj_base_date = base_date + \
#         adj_base_date.apply(lambda x: pd.Timedelta(x, unit='d'))
# #     delta = 4-base_date.isoweekday() if base_date.isoweekday()<=4 else 7+4-base_date.isoweekday()
# #     adj_base_date = base_date+timedelta(days=delta)

#     is_overdue = comparision_date > adj_base_date

# #     adj_base_date = adj_base_date.dt.strftime('%Y.%m.%d')
# #     adj_base_date = adj_base_date.strftime('%Y.%m.%d')

#     return is_overdue, adj_base_date


df_cur_conf[cur_conf_col] = df_cur_conf['Confirmed_Date'].astype('str') + \
    "_" + df_cur_conf['Confirmed_Quantity'].astype('str') + "/" \
    + (df_cur_conf['Order_Quantity'] -
       df_cur_conf['Delivery_Quantity']).astype('str')

# already confirmed but not taken by customers
df_forgoten = df_cur_conf[(df_cur_conf['Delivery_Quantity'] == 0) &
                          (df_cur_conf['Confirmed_Date'] < df_cur_conf['Created_on'].max())]

df_cur_conf = df_cur_conf.drop(columns=['Sold-To_Party',
                                        'Sales_document', 'SD_Document_Category', 'Customer_Reference',
                                        'Sales_Document_Item', 'Material', 'Material_text', 'Standard_Lead_Time', 'MRP_Controller',
                                        'Created_on', 'Order_Quantity', 'Delivery_Quantity'])


print(
    f'There are {df_part_delivered.shape[0]} lines partially delivered since last session')
# print('Here is a list of current confirmations')
# print(df_cur_conf.head())

if df_forgoten.shape[0]:
    print(f'There are {df_forgoten.shape[0]} lines in "forgotten" orders')
    print(df_forgoten.head())


def columns_to_drop(df, start_index=0, pattern='Confirmation_'):
    columns = df.columns

    drop_list = df.columns[start_index:]
    for stop_index, value in enumerate(columns):
        if pattern in value:
            drop_list = df.columns[start_index:stop_index]
            break

    return drop_list


df_conf, sheets = read_in_file(
    file=confirmation_file['file'], folder=file_folder, header_row=0)
# print (sheets)

# TODO multyprocess file read
# TODO confirmed for last Thursday - out of csope as this already in transit == delivered
# drop old empty columns from confirmation file 
if df_conf.size == 0:
    write_out_file(df=df_cur_conf,
              file_name=confirmation_file['file'], mode='w', sheet_name=cur_conf_col, if_sheet_exists='replace')
else:
    df_conf.set_index('item_index', drop=True, inplace=True)
#     df_conf.pop('changed')

    df_conf = df_conf.drop(columns=columns_to_drop(
        df_conf, start_index=1, pattern='Confirmation_'))
    df_conf.dropna(axis='columns', how='all', inplace=True)

# check if column with current date already exists and replace it
if cur_conf_col in df_conf.columns:
    print(
        f'\ncolumn {cur_conf_col} exists in file {confirmation_file["file"]}. Replacing it with new data')
    df_conf.pop(cur_conf_col)

def adj_requested_date(date_str):
    '''returns next thursday from given date'''
    parts = date_str if isinstance(
        date_str, date) else datetime.strptime(date_str, '%Y.%m.%d')
    delta = 4-parts.isoweekday() if parts.isoweekday() <= 4 else 7+4-parts.isoweekday()

#     if past:
#         delta -=7

    parts = parts+timedelta(days=delta)
    parts = parts.strftime('%Y.%m.%d')

    return parts

def focus_records(df):
    '''mark records of highest focus'''

    df_working = df.copy()
    df_working = df_working.drop_duplicates()

    insert_index = len(columns_to_drop(
        df_working, start_index=0, pattern='Confirmation_'))

    df_working['Focus'] = ''
#     df_working['Previous_confirmation'] = df_working[df_working.columns[insert_index+1]].apply(
#         lambda x: x if pd.isna(x) else x.split('_')[0])
    df_working[['Prev_conf_date', 'Prev_conf_qty']] = df_working[df_working.columns[insert_index+1]].apply(
        lambda x: pd.Series([None, 0]) if pd.isna(x) else pd.Series(x.split('_')))
    df_working['Prev_conf_qty'] = df_working['Prev_conf_qty'].apply(
        lambda x: int(x.split('/')[0]) if x else 0)
    df_working['Order_qty'] = df_working[df_working.columns[insert_index]].apply(
        lambda x: int(x.split('/')[1]) if x else 0)

    # False if latest confirmation is fully equal to previous date_qty/order_qty
    df_working['changed'] = df_working[df_working.columns[insert_index]
                                       ] != df_working[df_working.columns[insert_index+1]]

    df_working.reset_index(inplace=True, drop=True)

    unchanged_record = df_working[df_working['changed'] ==
                                  False].iloc[:, insert_index].drop_duplicates().values
    # all current confirmations

    # cur_confirmations = df_working.iloc[:,
    #                                     insert_index].drop_duplicates().values

    df_working.drop(df_working[df_working['changed'] == True
                               & df_working.iloc[:, insert_index].isin(unchanged_record)].index, inplace=True)
    df_working.drop(df_working[df_working['changed'] == True
                               & df_working.iloc[:, insert_index+1].isin(unchanged_record)].index, inplace=True)

    # if chanded both - dates and qty - select what to keep
    if df_working['Order_qty'].max() != df_working['Confirmed_Quantity'].sum():
        x = abs(df_working['Confirmed_Quantity'] -
                df_working['Prev_conf_qty']).min()
        df_working.drop(df_working[abs(
            df_working['Confirmed_Quantity'] - df_working['Prev_conf_qty']) > x].index, inplace=True)

    # orders without previous confirmation are considered 'not changed'
    df_working.loc[df_working.iloc[:, insert_index+1].isnull(),
                   'changed'] = False

    df_working['improved'] = df_working['Confirmed_Date'] < df_working['Prev_conf_date']
    df_working.fillna({'improved': True}, inplace=True)

    df_working['Adj_Requested_Date'] = df_working['Requested_Date'].apply(
        adj_requested_date)
    df_working['overdue'] = df_working['Confirmed_Date'] > df_working['Adj_Requested_Date']

    df_working['Next_delivery'] = adj_requested_date(date.today())
    df_working['Next_delivery'] = df_working['Confirmed_Date'] == df_working['Next_delivery']

    #     UBC Group
    df_working.loc[
        (df_working['Next_delivery'] == False)
        & (df_working['overdue'] == True)
        & (
            (df_working['Sold-To_Party_text'] == 'PSC "Ukpostach"')
            | (df_working['Sold-To_Party_text'] == 'LLC "Green Cool"')),
        ['Focus']] += 'UBC_overdue, '

    df_working.loc[
        (df_working['changed'] == True)
        & (df_working['improved'] == False)
        & (df_working['overdue'] == True)
        & (
            (df_working['Sold-To_Party_text'] == 'PSC "Ukpostach"')
            | (df_working['Sold-To_Party_text'] == 'LLC "Green Cool"')),
        ['Focus']] += 'UBC_changed, '
    
    #     All customers
    df_working.loc[
        (df_working['Next_delivery'] == False)
        & (df_working['overdue'] == True)
        & (df_working['Total_Stock'] > 0),
        ['Focus']] += 'overdue+available, '

    df_working.drop(labels=['Adj_Requested_Date',
                    'Next_delivery', 'Prev_conf_date', 'Prev_conf_qty'], axis=1, inplace=True)

    col = list(df_working.columns)
    col = col[:insert_index] + \
        col[len(df.columns):] + col[insert_index:len(df.columns)]
    df_working = df_working[col]

    return df_working

# check confirmation changes
items = df_cur_conf['item_index'].drop_duplicates().reset_index(drop=True)

# insert_index = len(columns_to_drop(df_cur_conf, start_index = 0, pattern='overdue'))
insert_index = len(columns_to_drop(
    df_cur_conf, start_index=0, pattern='Confirmation_'))

for item in items:

    try:
        # if there is a history on item
        ddf = df_cur_conf[df_cur_conf['item_index'] == item].join(
            df_conf.loc[[item], df_conf.columns[1:]], how='left', on='item_index')
    except:
        # if item is new in list
        ddf = df_cur_conf[df_cur_conf['item_index'] == item].join(
            pd.DataFrame(columns=df_conf.columns[1:], index=[item]), how='left', on='item_index')

    ddf = focus_records(ddf)

    if item == items[0]:
        df_conf_temp = ddf.copy()

    else:
        df_conf_temp = pd.concat([df_conf_temp, ddf], ignore_index=True)


# # for test purposes
# with pd.ExcelWriter('Deliveries\\temp.xlsx', mode='w') as writer:
#     df_conf_temp.to_excel(writer, sheet_name='test', index=False)
# print('Debug exit')
# os._exit(0)

df_conf = df_conf_temp
del (df_conf_temp)

# create grouping rule
group_col = df.columns.to_list()
my_dict = dict.fromkeys(group_col, 'max')
my_dict['Confirmed_Quantity'] = 'sum'
my_dict


def form_output_df(df1):

    print('filter not confirmed orders, created not earlier than 2 weeks ago')
    print('and requested not later than in 6 weeks from now\n')
    df2 = df1[(df1['Requested_Date'] <=
                (date.today() + timedelta(weeks=6)).strftime('%Y.%m.%d')) &
                (df1['Created_on'] <=
                (date.today() - timedelta(weeks=2)).strftime('%Y.%m.%d'))]
    df2 = df2[['PL', 'Sold-To_Party', 'Sold-To_Party_text', 'Sales_document',
                'Customer_Reference', 'Sales_Document_Item', 'Material',
                'Material_text', 'MRP_Controller', 'Created_on', 'Requested_Date',
                'Confirmed_Date', 'Order_Quantity', 'not_confirmed']]

    if len(df2) == 0:
        print("No unconfirmed orders found")

    print(df2.shape[0], " rows filtered\n")
    return df2


df1 = df.groupby('item_index', as_index=False).agg(my_dict)
df1['not_confirmed'] = df1['Order_Quantity'] - df1['Confirmed_Quantity']

# filter not confirmed lines only
df1 = df1[df1['not_confirmed'] > 0]

if len(df1):

    # filter_Rest = \
    #     'PL != "40" & Material != "080G5300" & Material != "080G5350" & Material != "080G5360" & Material != "080G5400"'

    # df_Rest = df1.query(filter_Rest)

    print('Total:', df1.shape)

    df_Rest_out = form_output_df(df1)

    if df_Rest_out.shape[0]:
        write_out_file(
            df=df_Rest_out, file_name=out_file['file'], mode='w', sheet_name='filter_Rest')
    else:
        print('Nothing to write to file')
else:
    print('No not confirmed orders found')

# save changes in delivery dates
write_out_file(df=df_conf, file_name=confirmation_file['file'],
          mode='a', sheet_name=cur_conf_col, if_sheet_exists='replace')

